#!/usr/bin/env python3
# coding: utf-8
"""
Fixy Data — Generador de Dashboard Offline
==========================================

Script self-contained que lee la carpeta "Base para dashboard offline" y
genera un dashboard HTML completo (index.html) listo para:
  - Compartir como archivo offline con la dirección
  - Publicar en https://javierbartaburu-dotcom.github.io/datafixy/

Datos consumidos:
  - Tablero Fixy 2026.xlsx
  - P&L_FIXY_2026_REAL.xlsx (Ventas, UB, EBITDA, NOPAT mes a mes + YTD + mix por UN)
  - Fulfillment 2026 (2).xlsx
  - Base de guías procesadas/consulta_guias_*.xls (Estado + COD)
  - Reportes SLA 2026/export_sla_*.csv (SLA + 1ra/2da/3ra visita + días)

KPIs principales:
  1. % Entregadas guías SIN contrarreembolso (COD=0) mes a mes + YTD
  2. % Entregadas guías CON contrarreembolso (COD>0) mes a mes + YTD
  3. % Cumplimiento del objetivo +70% YoY (vs volumen 2025) mes a mes + YTD
  4. % Entrega 1ra / 2da / 3ra visita con cantidades mes a mes + YTD
  5. P&L mensual real (Ene-Abr 2026) — Ventas / UB / Gastos / EBITDA / NOPAT
  6. Mix de ventas por unidad de negocio
  7. Punto de equilibrio
  8. SLA total, días, repactaciones, distribuidores, top clientes/zonas/servicios

Filtros: Vistage y Renata Arg Logística se excluyen de la vista comercial.

Uso:
    python generate_dashboard.py
    # opcional:
    FIXY_DATA_DIR="/ruta/a/carpeta" python generate_dashboard.py

Publicar:
    git add . && git commit -m "Update dashboard" && git push
"""
import os, sys, json, re, csv, collections, base64, getpass
from datetime import datetime

# ============================================================
# Configuración
# ============================================================
HOME = os.path.expanduser("~")
DEFAULT_BASE = os.path.join(HOME, "OneDrive", "Escritorio", "Base para dashboard offline")
if not os.path.isdir(DEFAULT_BASE):
    DEFAULT_BASE = os.path.join(HOME, "Escritorio", "Base para dashboard offline")
BASE = os.environ.get("FIXY_DATA_DIR", DEFAULT_BASE)
DASHBOARD_DIR = os.path.dirname(os.path.abspath(__file__))

print("=" * 60)
print("Fixy Data — Generador de Dashboard")
print("=" * 60)
print(f"Fuente: {BASE}")
print(f"Salida: {DASHBOARD_DIR}\n")

if not os.path.isdir(BASE):
    print(f"ERROR: no se encuentra la carpeta de datos {BASE}")
    print("Definí FIXY_DATA_DIR como variable de entorno apuntando a la carpeta correcta.")
    sys.exit(1)

csv.field_size_limit(10_000_000)

SLA_DIR = os.path.join(BASE, "Reportes SLA 2026")
GUIAS_DIR = os.path.join(BASE, "Base de guías procesadas")
PNL_FILE = os.path.join(BASE, "P&L_FIXY_2026_REAL.xlsx")

# Volúmenes 2025 (del Tablero Fixy 2026, hoja 2025) para target +70%
TABLERO_2025 = {1:24606,2:16649,3:17328,4:20604,5:24525,6:27649,7:40948,
                8:26637,9:29626,10:27295,11:29863,12:38113}
DIAS_MES = {1:31,2:28,3:31,4:30,5:31,6:30,7:31,8:31,9:30,10:31,11:30,12:31}
DIAS_PARCIAL = 22

EXCLUIR_RE = re.compile(r'\b(vistage|renata)\b', re.I)

MESES_NUM = {'enero':1,'febrero':2,'marzo':3,'abril':4,'mayo':5,
             'junio':6,'julio':7,'agosto':8,'septiembre':9,'octubre':10,
             'noviembre':11,'diciembre':12}
MESES_NOM = {v:k.title() for k,v in MESES_NUM.items()}

# ============================================================
# Detección de archivos
# ============================================================
def find_sla_files():
    out = {}
    if not os.path.isdir(SLA_DIR): return out
    for f in os.listdir(SLA_DIR):
        low = f.lower()
        if not low.endswith('.csv') or not low.startswith('export_sla'): continue
        for nombre, num in MESES_NUM.items():
            if nombre in low:
                parcial = 'parcial' in low
                out[num] = (f, (MESES_NOM[num]+'*' if parcial else MESES_NOM[num]), parcial)
                break
    return out

def find_guia_files():
    out = {}
    if not os.path.isdir(GUIAS_DIR): return out
    for f in os.listdir(GUIAS_DIR):
        if not f.lower().endswith('.xls'): continue
        m = re.search(r'(0[1-9]|1[0-2])\.2026', f)
        if m: out[int(m.group(1))] = f
    return out

SLA_FILES = find_sla_files()
GUIA_FILES = find_guia_files()
MESES = sorted(SLA_FILES.keys())
NOMBRES = [SLA_FILES[m][1] for m in MESES]
print(f"SLA encontrados: {MESES}")
print(f"Guías encontrados: {sorted(GUIA_FILES.keys())}\n")
if not MESES:
    print("ERROR: no se encontraron exports SLA en", SLA_DIR); sys.exit(1)

# ============================================================
# Helpers
# ============================================================
def pf(s):
    if not s: return 0.0
    s = str(s).strip().strip('"').replace(",", ".")
    if not s: return 0.0
    try: return float(s)
    except: return 0.0

def pct_q(sl, q):
    if not sl: return 0
    k = (len(sl)-1) * q; f = int(k); c = min(f+1, len(sl)-1)
    return sl[f] if f==c else sl[f] + (sl[c]-sl[f]) * (k-f)

def is_entregada(estado):
    e = (estado or '').upper()
    return 'ENTREGADA' in e or e.startswith('POD')

# ============================================================
# 1) SLA mensual
# ============================================================
print("[1/4] Procesando SLA mensual...")
month_sla = {}
clientes_m = collections.defaultdict(collections.Counter)
sucursales_m = collections.defaultdict(collections.Counter)
zonas_m = collections.defaultdict(collections.Counter)
servicios_m = collections.defaultdict(collections.Counter)

for mes_n in MESES:
    fname, mes_nombre, _ = SLA_FILES[mes_n]
    path = os.path.join(SLA_DIR, fname)
    with open(path, encoding='utf-8', errors='ignore') as fh:
        first = fh.readline()
    delim = '|' if '|' in first[:200] else ';'

    n=0; ne=0; ndc=0; ndrg=0; nel=0; nint=0; no=0
    nv1=nv2=nv3=0; nrep=0
    nev1=nev2=nev3=nesv=0
    dias=[]; buckets=collections.Counter()

    with open(path, encoding='utf-8', errors='ignore', newline='') as fh:
        rdr = csv.reader(fh, delimiter=delim, quotechar='"')
        H = {h.strip().strip('"'): i for i,h in enumerate(next(rdr))}
        max_idx = max(H.values())
        for row in rdr:
            if len(row) <= max_idx: continue
            n += 1
            cliente = (row[H["CLIENTE"]] or "").strip().strip('"')
            sucursal = (row[H["SUCURSAL CLIENTE"]] or "").strip().strip('"')
            servicio = (row[H["SERVICIO"]] or "").strip().strip('"')
            grupo = (row[H["GRUPO"]] or "").strip().strip('"').upper()
            sub_dest = (row[H["SUB ZONA DESTINO"]] or "").strip().strip('"').upper() or "SIN ZONA"
            v1 = bool((row[H["FECHA PRIMERA VISITA"]] or "").strip().strip('"'))
            v2 = bool((row[H["FECHA SEGUNDA VISITA"]] or "").strip().strip('"'))
            v3 = bool((row[H["FECHA TERCER VISITA"]] or "").strip().strip('"'))
            if v1: nv1 += 1
            if v2: nv2 += 1
            if v3: nv3 += 1
            frep = (row[H["FECHA REPACTACION"]] or "").strip().strip('"')
            if frep and frep != "0": nrep += 1
            if "ENTREGADA" in grupo: ne += 1
            elif "DEVOLUCION CON COBRO" in grupo: ndc += 1
            elif "DEVOLUCION" in grupo and "RENDICION GENERADA" in grupo: ndrg += 1
            elif "ELIMINADO" in grupo: nel += 1
            elif "INTERIOR PAIS" in grupo: nint += 1
            else: no += 1
            if "ENTREGADA" in grupo:
                if v1 and not v2: nev1 += 1
                elif v2 and not v3: nev2 += 1
                elif v3: nev3 += 1
                else: nesv += 1
            td = pf(row[H["TOTAL DIAS"]])
            if 0 <= td <= 365:
                dias.append(td)
                if td <= 1: buckets["0-1"] += 1
                elif td <= 3: buckets["1-3"] += 1
                elif td <= 7: buckets["3-7"] += 1
                elif td <= 14: buckets["7-14"] += 1
                elif td <= 30: buckets["14-30"] += 1
                else: buckets["30+"] += 1
            zonas_m[mes_n][sub_dest] += 1
            servicios_m[mes_n][servicio] += 1
            if cliente: clientes_m[mes_n][cliente] += 1
            if cliente and sucursal: sucursales_m[mes_n][f"{cliente}||{sucursal}"] += 1

    dias.sort()
    et = nev1 + nev2 + nev3 + nesv
    month_sla[mes_n] = dict(
        nombre=mes_nombre, total=n, entregadas=ne, dev_cobro=ndc, dev_rg=ndrg,
        eliminadas=nel, interior=nint, otro=no,
        v1=nv1, v2=nv2, v3=nv3, repactadas=nrep,
        ent_v1=nev1, ent_v2=nev2, ent_v3plus=nev3, ent_sin_visita=nesv,
        sla_correcto=ne/n if n else 0,
        sla_anterior=(ne+ndc)/n if n else 0,
        pct_1visita=nev1/et if et else 0,
        dias_avg=sum(dias)/len(dias) if dias else 0,
        dias_p50=pct_q(dias,0.5), dias_p90=pct_q(dias,0.9),
        dias_max=max(dias) if dias else 0,
        dias_buckets=dict(buckets),
    )
    print(f"  {mes_nombre}: N={n}, Ent={ne}, SLA={ne/n*100:.1f}%")

# ============================================================
# 2) consulta_guias — COD, cumplimiento, comercial filtrado
# ============================================================
print("\n[2/4] Procesando consulta_guias...")
month_cv = {}
clientes_filt_m = collections.defaultdict(collections.Counter)
top_dist_yt = collections.Counter()

for mes_n in MESES:
    if mes_n not in GUIA_FILES:
        month_cv[mes_n] = {'total':0,'entregadas':0,'sla_pct':0,
            'cod0_total':0,'cod0_entregadas':0,'cod0_sla_pct':0,
            'codpos_total':0,'codpos_entregadas':0,'codpos_sla_pct':0,
            'target_2026':0,'target_prorrateado':0,'cumplimiento_pct':0,
            'filas':0,'n_dist':0,'importe_sum':0,'contrareembolso_sum':0,
            'n_excluidos':0,'total_filtered':0,'entregadas_filtered':0}
        continue
    fn = GUIA_FILES[mes_n]
    path = os.path.join(GUIAS_DIR, fn)
    txt = open(path, encoding='utf-8', errors='ignore').read()
    rows = re.findall(r'<tr>(?!\s*<th)(.*?)</tr>', txt, re.S)

    n=0; ne=0; n_cod0=0; n_cod0_e=0; n_codp=0; n_codp_e=0
    n_filt_total=0; n_filt_ent=0; n_exclu=0
    dist_rec = collections.Counter()
    importe_sum=0.0; cr_sum=0.0

    for row in rows:
        cells = re.findall(r'<td>(.*?)</td>', row, re.S)
        if len(cells) < 38: continue
        cells = [c.strip() for c in cells]
        estado = cells[4]
        cliente = cells[2]
        empresa = cells[18] if len(cells)>18 else ""
        importe = pf(cells[16] if len(cells)>16 else "")
        cr_val = pf(cells[37] if len(cells)>37 else "")
        n += 1
        es_e = is_entregada(estado)
        if es_e: ne += 1
        if cr_val == 0:
            n_cod0 += 1
            if es_e: n_cod0_e += 1
        else:
            n_codp += 1
            if es_e: n_codp_e += 1
        if len(cells)>25 and cells[25]: dist_rec[cells[25]] += 1
        importe_sum += importe
        if cr_val > 0: cr_sum += cr_val
        cli_str = (cliente or '') + ' ' + (empresa or '')
        if EXCLUIR_RE.search(cli_str):
            n_exclu += 1
            continue
        n_filt_total += 1
        if es_e: n_filt_ent += 1
        if cliente: clientes_filt_m[mes_n][cliente] += 1

    target_2026 = TABLERO_2025.get(mes_n, 0) * 1.70
    parcial = SLA_FILES[mes_n][2]
    target_prorrateado = target_2026 * DIAS_PARCIAL / DIAS_MES[mes_n] if parcial else target_2026
    month_cv[mes_n] = {
        'total':n,'entregadas':ne,'sla_pct':ne/n if n else 0,
        'cod0_total':n_cod0,'cod0_entregadas':n_cod0_e,'cod0_sla_pct':n_cod0_e/n_cod0 if n_cod0 else 0,
        'codpos_total':n_codp,'codpos_entregadas':n_codp_e,'codpos_sla_pct':n_codp_e/n_codp if n_codp else 0,
        'target_2026':target_2026,'target_prorrateado':target_prorrateado,
        'cumplimiento_pct':n/target_prorrateado if target_prorrateado else 0,
        'filas':n,'n_dist':len(dist_rec),
        'importe_sum':importe_sum,'contrareembolso_sum':cr_sum,
        'n_excluidos':n_exclu,'total_filtered':n_filt_total,'entregadas_filtered':n_filt_ent,
    }
    for name, c in dist_rec.most_common(15):
        top_dist_yt[name] += c
    print(f"  {SLA_FILES[mes_n][1]}: N={n}, COD=0 SLA={month_cv[mes_n]['cod0_sla_pct']*100:.1f}%, Cumpl={month_cv[mes_n]['cumplimiento_pct']*100:.1f}%")

# ============================================================
# 3) P&L Real — Mensual + YTD + mix UN + punto equilibrio
# ============================================================
print("\n[3/4] Procesando P&L Real 2026...")
pnl_monthly = {}     # m -> dict con ventas, ub, gastos, ebitda, nopat, %
pnl_ytd = {}          # dict YTD a Abril
pnl_mix = []          # mix por unidad de negocio
pnl_breakeven = {}    # punto de equilibrio del mes seleccionado en Dashboard

try:
    import openpyxl
    wb = openpyxl.load_workbook(PNL_FILE, read_only=True, data_only=True)

    # ---- Dashboard sheet: KPIs anuales, mix, punto equilibrio ----
    ws = wb['Dashboard']
    rows = list(ws.iter_rows(values_only=True))

    def safe(r, c):
        try:
            v = rows[r][c]
            return v if isinstance(v, (int, float)) else 0
        except: return 0

    pnl_ytd = {
        'ventas': safe(6, 1),
        'utilidad_bruta': safe(6, 4),
        'ebitda': safe(6, 7),
        'margen_ebitda': safe(6, 10),
        'nopat': safe(10, 1),
        'margen_nopat': safe(10, 4),
        'cv_sobre_ventas': safe(10, 7),
        'costos_fijos': safe(10, 10),
        'margen_bruto': safe(6, 4) / safe(6, 1) if safe(6, 1) else 0,
    }
    print(f"  YTD: Ventas={pnl_ytd['ventas']/1e6:.1f}M, EBITDA={pnl_ytd['ebitda']/1e6:.1f}M")

    # Mix UN (R16-19)
    for r in [16, 17, 18, 19]:
        if r < len(rows):
            un = str(rows[r][1] or '').strip()
            ventas = safe(r, 4)
            mix = safe(r, 7)
            if un and ventas:
                pnl_mix.append({'un':un, 'ventas':ventas, 'mix':mix})

    # Punto equilibrio (R23-27 col 8)
    pnl_breakeven = {
        'mes_seleccionado': str(rows[23][8] or '') if len(rows)>23 else '',
        'ventas_reales': safe(24, 8),
        'pe_sin_iva': safe(25, 8),
        'pe_con_iva': safe(26, 8),
        'excedente': safe(27, 8),
    }

    # ---- Ventas sheet: envíos CrossDocking ----
    try:
        ws_v = wb['Ventas']
        rows_v = list(ws_v.iter_rows(values_only=True))
        # R6 = CrossDocking quantities; columnas 1..12 = ene..dic
        envios_crossdocking = []
        for c in range(1, 6):  # ene-may
            v = rows_v[6][c] if len(rows_v)>6 and c < len(rows_v[6]) else None
            envios_crossdocking.append(v if isinstance(v,(int,float)) and v else 0)
        pnl_envios = envios_crossdocking
        print(f"  Envíos CrossDocking Ene-May: {envios_crossdocking}")
    except Exception as e:
        print(f"  WARN — no se pudo leer hoja Ventas: {e}")
        pnl_envios = [34201, 28519, 26917, 33053, 0]

    # ---- FIXY P&L 2026: mensual ----
    ws = wb['FIXY P&L 2026']
    rows = list(ws.iter_rows(values_only=True))
    # columnas 14, 15, 16, 17 = Ene/Feb/Mar/Abr 2026
    cols_2026 = {1:14, 2:15, 3:16, 4:17}  # ago mostraría 18 etc, pero por ahora YTD a Abril
    def gv(r, c):
        try:
            v = rows[r][c]
            return v if isinstance(v, (int, float)) else 0
        except: return 0

    for mes_n in [1,2,3,4]:
        c = cols_2026[mes_n]
        ventas = gv(6, c)
        cv = gv(12, c)
        ub = gv(13, c)
        margen_b = gv(14, c)
        gastos = gv(35, c)
        ebitda = gv(36, c)
        margen_e = gv(37, c)
        nopat = gv(42, c)
        margen_n = gv(43, c)
        pnl_monthly[mes_n] = dict(
            ventas=ventas, cv=cv, ub=ub, margen_bruto=margen_b,
            gastos=gastos, ebitda=ebitda, margen_ebitda=margen_e,
            nopat=nopat, margen_nopat=margen_n,
        )

    wb.close()
    print(f"  Mensual: 4 meses (Ene-Abr 2026)")
    print(f"  Mix UN: {len(pnl_mix)} unidades")
    print(f"  Punto equilibrio mes seleccionado: {pnl_breakeven['mes_seleccionado']}")
except Exception as e:
    print(f"  WARN — no se pudo leer P&L: {e}")
    pnl_ytd = {'ventas':0,'utilidad_bruta':0,'ebitda':0,'margen_ebitda':0,'nopat':0,
               'margen_nopat':0,'cv_sobre_ventas':0,'costos_fijos':0,'margen_bruto':0}

# ============================================================
# 4) Consolidación
# ============================================================
print("\n[4/4] Consolidando KPIs...")
total_y_sla = sum(month_sla[m]['total'] for m in MESES)
ent_y_sla = sum(month_sla[m]['entregadas'] for m in MESES)
sla_y = ent_y_sla/total_y_sla if total_y_sla else 0
total_y_cv = sum(month_cv[m]['total'] for m in MESES)
cod0_y = sum(month_cv[m]['cod0_total'] for m in MESES)
cod0_ent_y = sum(month_cv[m]['cod0_entregadas'] for m in MESES)
codpos_y = sum(month_cv[m]['codpos_total'] for m in MESES)
codpos_ent_y = sum(month_cv[m]['codpos_entregadas'] for m in MESES)
target_y = sum(month_cv[m]['target_prorrateado'] for m in MESES)
cumpl_y = total_y_cv/target_y if target_y else 0
imp_y = sum(month_cv[m]['importe_sum'] for m in MESES)
cod_in_y = sum(month_cv[m]['contrareembolso_sum']*0.01 for m in MESES)
ev1y = sum(month_sla[m]['ent_v1'] for m in MESES)
ev2y = sum(month_sla[m]['ent_v2'] for m in MESES)
ev3y = sum(month_sla[m]['ent_v3plus'] for m in MESES)
esvy = sum(month_sla[m]['ent_sin_visita'] for m in MESES)
evty = ev1y+ev2y+ev3y+esvy

totC_filt = collections.Counter()
for m in MESES:
    for c,n in clientes_filt_m[m].items(): totC_filt[c] += n
top15_filt = totC_filt.most_common(15)
serie_clientes = {c:[clientes_filt_m[m].get(c,0) for m in MESES] for c,_ in top15_filt}

# Concentración Top 3 — por mes y YTD (sobre vista comercial filtrada)
top3_volumen_por_mes = []   # volumen absoluto top 3 del mes
top3_pct_por_mes = []       # % sobre total del mes
top3_nombres_por_mes = []   # 3 clientes del mes
total_filt_por_mes = [sum(clientes_filt_m[m].values()) for m in MESES]
for m in MESES:
    top3_mes = clientes_filt_m[m].most_common(3)
    vol3 = sum(v for _,v in top3_mes)
    tot = sum(clientes_filt_m[m].values())
    top3_volumen_por_mes.append(vol3)
    top3_pct_por_mes.append(round(vol3/tot*100,1) if tot else 0)
    top3_nombres_por_mes.append([{'cliente':c,'guias':v} for c,v in top3_mes])
# YTD: top 3 del agregado
top3_ytd = totC_filt.most_common(3)
top3_ytd_volumen = sum(v for _,v in top3_ytd)
top3_ytd_total = sum(totC_filt.values())
top3_ytd_pct = top3_ytd_volumen/top3_ytd_total*100 if top3_ytd_total else 0
# Top 10 share YTD
top10_ytd_share = sum(v for _,v in totC_filt.most_common(10))/top3_ytd_total*100 if top3_ytd_total else 0

sucC = collections.Counter()
for m in MESES:
    for cs,n in sucursales_m[m].items(): sucC[cs] += n

zonas_total = collections.Counter()
zonas_ent_real = collections.defaultdict(float)
for m in MESES:
    prop = month_sla[m]['entregadas']/month_sla[m]['total'] if month_sla[m]['total'] else 0
    for z,c in zonas_m[m].items():
        zonas_total[z] += c
        zonas_ent_real[z] += c*prop
top15_zonas = zonas_total.most_common(15)
top8z = [z for z,_ in zonas_total.most_common(8)]
heatmap = [{'zona':z, 'data':[zonas_m[m].get(z,0) for m in MESES]} for z in top8z]

servC = collections.Counter()
for m in MESES:
    for s,c in servicios_m[m].items(): servC[s] += c
top10_serv = servC.most_common(10)
serv_mes = {s:[servicios_m[m].get(s,0) for m in MESES] for s,_ in top10_serv}

churn = []
for i in range(len(MESES)-1):
    a,b = MESES[i], MESES[i+1]
    sA,sB = set(clientes_m[a].keys()), set(clientes_m[b].keys())
    churn.append({'from':NOMBRES[i],'to':NOMBRES[i+1],
                  'continuan':len(sA&sB),'bajas':len(sA-sB),'altas':len(sB-sA)})

# P&L mensual arrays (4 meses)
pnl_meses = ['Enero','Febrero','Marzo','Abril']
pnl_arrays = {
    'ventas': [pnl_monthly.get(m,{}).get('ventas',0) for m in [1,2,3,4]],
    'ub':     [pnl_monthly.get(m,{}).get('ub',0) for m in [1,2,3,4]],
    'gastos': [abs(pnl_monthly.get(m,{}).get('gastos',0)) for m in [1,2,3,4]],
    'ebitda': [pnl_monthly.get(m,{}).get('ebitda',0) for m in [1,2,3,4]],
    'nopat':  [pnl_monthly.get(m,{}).get('nopat',0) for m in [1,2,3,4]],
    'margen_bruto':  [pnl_monthly.get(m,{}).get('margen_bruto',0)*100 for m in [1,2,3,4]],
    'margen_ebitda': [pnl_monthly.get(m,{}).get('margen_ebitda',0)*100 for m in [1,2,3,4]],
    'margen_nopat':  [pnl_monthly.get(m,{}).get('margen_nopat',0)*100 for m in [1,2,3,4]],
    'cv': [abs(pnl_monthly.get(m,{}).get('cv',0)) for m in [1,2,3,4]],
    'envios_cd': [int(round(v)) for v in (pnl_envios if 'pnl_envios' in dir() else [34201,28519,26917,33053])][:4],
}

# Calcular promedio MoM histórico de envíos
try:
    e = pnl_arrays['envios_cd']
    mom_rates = [(e[i+1]/e[i] - 1) for i in range(len(e)-1) if e[i]>0 and e[i+1]>0]
    mom_avg = sum(mom_rates)/len(mom_rates) if mom_rates else 0
    pnl_arrays['mom_envios_avg'] = round(mom_avg*100, 2)
except: pnl_arrays['mom_envios_avg'] = 0

# Regresión lineal simple (Y = a + b·X) para CF y CV
def linreg(xs, ys):
    n = len(xs)
    if n < 2 or all(x==0 for x in xs): return {'a':0,'b':0,'r2':0}
    sx = sum(xs); sy = sum(ys); sxx = sum(x*x for x in xs); sxy = sum(x*y for x,y in zip(xs,ys))
    den = n*sxx - sx*sx
    if den == 0: return {'a':0,'b':0,'r2':0}
    b = (n*sxy - sx*sy)/den
    a = (sy - b*sx)/n
    # R²
    yhat = [a + b*x for x in xs]
    ybar = sy/n
    ss_tot = sum((y-ybar)**2 for y in ys)
    ss_res = sum((y-yh)**2 for y,yh in zip(ys,yhat))
    r2 = 1 - ss_res/ss_tot if ss_tot>0 else 0
    return {'a':round(a,2),'b':round(b,2),'r2':round(r2,3)}

reg_cf = linreg(pnl_arrays['envios_cd'], pnl_arrays['gastos'])
reg_cv = linreg(pnl_arrays['envios_cd'], pnl_arrays['cv'])
pnl_arrays['reg_cf'] = reg_cf
pnl_arrays['reg_cv'] = reg_cv

data = {
    'meses':NOMBRES, 'pnl_meses':pnl_meses,
    # Volumen y SLA
    'volumen':[month_sla[m]['total'] for m in MESES],
    'volumen_courier':[month_cv[m]['total'] for m in MESES],
    'sla_correcto':[round(month_sla[m]['sla_correcto']*100,1) for m in MESES],
    'sla_anterior':[round(month_sla[m]['sla_anterior']*100,1) for m in MESES],
    'entregadas':[month_sla[m]['entregadas'] for m in MESES],
    'dev_cobro':[month_sla[m]['dev_cobro'] for m in MESES],
    'dev_rg':[month_sla[m]['dev_rg'] for m in MESES],
    'eliminadas':[month_sla[m]['eliminadas'] for m in MESES],
    'otro':[month_sla[m]['otro'] for m in MESES],
    # COD breakdown
    'cod0_total':[month_cv[m]['cod0_total'] for m in MESES],
    'cod0_entregadas':[month_cv[m]['cod0_entregadas'] for m in MESES],
    'cod0_sla_pct':[round(month_cv[m]['cod0_sla_pct']*100,1) for m in MESES],
    'codpos_total':[month_cv[m]['codpos_total'] for m in MESES],
    'codpos_entregadas':[month_cv[m]['codpos_entregadas'] for m in MESES],
    'codpos_sla_pct':[round(month_cv[m]['codpos_sla_pct']*100,1) for m in MESES],
    'sla_total_courier':[round(month_cv[m]['sla_pct']*100,1) for m in MESES],
    'total_courier':[month_cv[m]['total'] for m in MESES],
    # Cumplimiento
    'target_2026':[round(month_cv[m]['target_2026']) for m in MESES],
    'target_prorrateado':[round(month_cv[m]['target_prorrateado']) for m in MESES],
    'cumplimiento_pct':[round(month_cv[m]['cumplimiento_pct']*100,1) for m in MESES],
    # Visitas
    'ent_v1':[month_sla[m]['ent_v1'] for m in MESES],
    'ent_v2':[month_sla[m]['ent_v2'] for m in MESES],
    'ent_v3plus':[month_sla[m]['ent_v3plus'] for m in MESES],
    'ent_sin_v':[month_sla[m]['ent_sin_visita'] for m in MESES],
    'pct_v1':[round(month_sla[m]['pct_1visita']*100,1) for m in MESES],
    # Dias
    'dias_avg':[round(month_sla[m]['dias_avg'],2) for m in MESES],
    'dias_p50':[month_sla[m]['dias_p50'] for m in MESES],
    'dias_p90':[month_sla[m]['dias_p90'] for m in MESES],
    'dias_max':[month_sla[m]['dias_max'] for m in MESES],
    'dias_buckets':[month_sla[m]['dias_buckets'] for m in MESES],
    'repactadas':[month_sla[m]['repactadas'] for m in MESES],
    'repact_pct':[round(month_sla[m]['repactadas']/month_sla[m]['total']*100,1) for m in MESES],
    # Distribuidores
    'distribuidores':[month_cv[m]['n_dist'] for m in MESES],
    'top_distribuidores':[{'n':n,'g':g} for n,g in top_dist_yt.most_common(10)],
    # Ingresos
    'importe_flete':[month_cv[m]['importe_sum'] for m in MESES],
    'ingreso_cod':[month_cv[m]['contrareembolso_sum']*0.01 for m in MESES],
    # Comercial filtrado
    'top_clientes_filt':[{'cliente':c,'guias':v,'share':round(v/sum(totC_filt.values())*100,2),'serie':serie_clientes[c]} for c,v in top15_filt],
    'top_sucursales':[{'sucursal':cs.replace('||',' | '),'guias':v} for cs,v in sucC.most_common(15)],
    'top_zonas':[{'zona':z,'guias':v,'sla':round(zonas_ent_real[z]/v*100,1) if v else 0,'share':round(v/sum(zonas_total.values())*100,2)} for z,v in top15_zonas],
    'top_servicios':[{'servicio':s,'guias':v,'share':round(v/sum(servC.values())*100,2),'serie':serv_mes[s]} for s,v in top10_serv],
    'churn':churn,'heatmap':heatmap,
    # P&L
    'pnl_ytd':pnl_ytd,
    'pnl_mix':pnl_mix,
    'pnl_breakeven':pnl_breakeven,
    'pnl_arrays':pnl_arrays,
    # YTD
    'ytd_volumen':total_y_sla,'ytd_entregadas':ent_y_sla,'ytd_sla':round(sla_y*100,1),
    'ytd_cod0_sla':round(cod0_ent_y/cod0_y*100,1) if cod0_y else 0,
    'ytd_cod0_total':cod0_y,'ytd_cod0_ent':cod0_ent_y,
    'ytd_codpos_sla':round(codpos_ent_y/codpos_y*100,1) if codpos_y else 0,
    'ytd_codpos_total':codpos_y,
    'ytd_cumplimiento':round(cumpl_y*100,1),
    'ytd_target':round(target_y),'ytd_actual_courier':total_y_cv,
    'ytd_pct_v1':round(ev1y/evty*100,1) if evty else 0,
    'ytd_pct_v2':round(ev2y/evty*100,1) if evty else 0,
    'ytd_pct_v3':round(ev3y/evty*100,1) if evty else 0,
    'ytd_total_entregadas':evty,
    'ytd_dias_avg':round(sum(month_sla[m]['dias_avg'] for m in MESES)/len(MESES),2),
    'ytd_repact_pct':round(sum(month_sla[m]['repactadas'] for m in MESES)/total_y_sla*100,1) if total_y_sla else 0,
    'ytd_ingreso_total':imp_y+cod_in_y,'ytd_ingreso_flete':imp_y,'ytd_ingreso_cod':cod_in_y,
    'ytd_ticket_promedio':(imp_y+cod_in_y)/total_y_sla if total_y_sla else 0,
    'n_excluidos_ytd':sum(month_cv[m]['n_excluidos'] for m in MESES),
    # Concentración Top 3 / Top 10
    'top3_volumen_por_mes': top3_volumen_por_mes,
    'top3_pct_por_mes': top3_pct_por_mes,
    'top3_nombres_por_mes': top3_nombres_por_mes,
    'total_filt_por_mes': total_filt_por_mes,
    'top3_ytd_volumen': top3_ytd_volumen,
    'top3_ytd_total': top3_ytd_total,
    'top3_ytd_pct': round(top3_ytd_pct,1),
    'top3_ytd_nombres': [{'cliente':c,'guias':v} for c,v in top3_ytd],
    'top10_ytd_share': round(top10_ytd_share,1),
    # Año disponibles (para filtro)
    'anios_disponibles': [2026],
    # Metadata
    'fecha_corte':'22 de Mayo de 2026',
    'fecha_generado':datetime.now().strftime('%Y-%m-%d %H:%M'),
    'github_url':'https://javierbartaburu-dotcom.github.io/datafixy/',
    'repo_url':'https://github.com/javierbartaburu-dotcom/datafixy',
}

# Guardar data.json (sin encriptar — útil para auditoría local; gitignorable)
DATA_JSON = json.dumps(data, ensure_ascii=False)
data_path = os.path.join(DASHBOARD_DIR, "data.json")
with open(data_path, 'w', encoding='utf-8') as f:
    f.write(DATA_JSON)
print(f"\nOK data.json — {os.path.getsize(data_path)} bytes")

# ============================================================
# Encriptación AES-256-GCM para el dashboard público
# ============================================================
# La password puede setearse vía:
#   1) variable de entorno FIXY_DASHBOARD_PASSWORD
#   2) prompt interactivo si no está seteada
#   3) default "fixy2026" (con warning)
password = os.environ.get("FIXY_DASHBOARD_PASSWORD")
if not password:
    try:
        if sys.stdin.isatty():
            password = getpass.getpass("Password para el dashboard (Enter para default 'fixy2026'): ").strip()
    except: password = None
if not password:
    password = "fixy2026"
    print("WARN — usando password por defecto 'fixy2026'. Definí FIXY_DASHBOARD_PASSWORD para cambiarla.")

try:
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
    from cryptography.hazmat.primitives.ciphers.aead import AESGCM
    from cryptography.hazmat.primitives import hashes as _hashes
    salt = os.urandom(16)
    kdf = PBKDF2HMAC(algorithm=_hashes.SHA256(), length=32, salt=salt, iterations=200000)
    key = kdf.derive(password.encode('utf-8'))
    nonce = os.urandom(12)
    aesgcm = AESGCM(key)
    ct = aesgcm.encrypt(nonce, DATA_JSON.encode('utf-8'), None)
    encrypted_payload = {
        'v': 1,
        'kdf': 'PBKDF2-SHA256',
        'iter': 200000,
        'salt': base64.b64encode(salt).decode(),
        'nonce': base64.b64encode(nonce).decode(),
        'ct': base64.b64encode(ct).decode(),
    }
    ENC_BLOB = json.dumps(encrypted_payload)
    print(f"OK datos encriptados con AES-256-GCM (PBKDF2 200k iter) — {len(ENC_BLOB)} bytes")
    USE_ENC = True
except Exception as e:
    print(f"WARN — no se pudo encriptar (cryptography no disponible): {e}")
    print("       el dashboard saldrá sin password — instalá 'cryptography' con: pip install cryptography")
    ENC_BLOB = '{}'
    USE_ENC = False

# ============================================================
# 5) Generar HTML (template embebido)
# ============================================================
print("\n[5/5] Generando HTML...")
# El template HTML está en archivo separado dashboard_template.html en la misma carpeta.
# Si no existe, se reescribe sólo el bloque <script id="data"> del index.html actual.
template_path = os.path.join(DASHBOARD_DIR, "dashboard_template.html")
index_path = os.path.join(DASHBOARD_DIR, "index.html")
print(f"  template: {template_path} exists={os.path.isfile(template_path)}")
print(f"  index:    {index_path} exists={os.path.isfile(index_path)}")

if os.path.isfile(template_path):
    with open(template_path, encoding='utf-8') as f:
        html = f.read()
    if '__ENC_BLOB__' in html and USE_ENC:
        html = html.replace('__ENC_BLOB__', ENC_BLOB)
        html = html.replace('__DATA_JSON__', '{}')
        html = html.replace('__USE_ENC__', 'true')
    else:
        html = html.replace('__ENC_BLOB__', '{}')
        html = html.replace('__DATA_JSON__', DATA_JSON)
        html = html.replace('__USE_ENC__', 'false')
    html = html.replace('__FECHA_GEN__', data['fecha_generado'])
    html = html.replace('__CORTE__', data['fecha_corte'])
    html = html.replace('__GH_URL__', data['github_url'])
    html = html.replace('__GH_URL_TXT__', data['github_url'].replace('https://',''))
    html = html.replace('__REPO_URL__', data['repo_url'])
    with open(index_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"OK index.html generado desde template - {os.path.getsize(index_path)} bytes")
elif os.path.isfile(index_path):
    txt = open(index_path, encoding='utf-8').read()
    new_txt = re.sub(r'(<script id="data" type="application/json">)[\s\S]*?(</script>)', lambda m: m.group(1) + '\n' + DATA_JSON + '\n' + m.group(2), txt, count=1)
    new_txt = re.sub(r'Ultima actualizacion: [0-9\- :]+', f'Ultima actualizacion: {data["fecha_generado"]}', new_txt)
    with open(index_path, 'w', encoding='utf-8') as f:
        f.write(new_txt)
    print(f"OK index.html actualizado - {os.path.getsize(index_path)} bytes")
else:
    print("WARN: ni template ni index existen.")

print("\n" + "="*60)
print("LISTO. Para publicar en GitHub Pages:")
print("  cd " + DASHBOARD_DIR)
print("  git add . && git commit -m 'Update dashboard' && git push")
print("\nSitio: https://javierbartaburu-dotcom.github.io/datafixy/")
print("="*60)
